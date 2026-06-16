"""
xlsx_lint.py — OBJECTIVE-DEFECT linter for .xlsx workbooks.

Same scope philosophy as pptx_lint.py: flag ONLY objective, unambiguous mistakes
that a strictly MORE capable model would also never want. Two tests per check:
  (1) OBJECTIVE  — wrong regardless of taste/intent.
  (2) NO-SHACKLE — a smarter model would still always avoid it.

DELIBERATELY NOT CHECKED (these need the user's ORIGINAL as a baseline, so they are
PROCESS-discipline, not lintable from the file alone — see GUIDELINE.md C1/C3):
  "Claude added a cell color that breaks our internal convention" — a color is not
  objectively wrong; it's non-conformance to a baseline we cannot see here. Also out:
  palette/font/number-format taste, column-width aesthetics, density.

What IS objective and lintable:
  XL-FORMULA-ERR  cells with error results / #REF! baked into a formula      ERROR
  XL-MERGE-OVERLAP overlapping merged ranges                                  ERROR
  XL-BORDER-GAP   a data block whose border outline is PARTIAL (case 4)       WARN
  XL-NUM-AS-TEXT  a number stored as text inside a numeric column             WARN
  XL-ARTIFACT     literal markdown / leftover placeholder text in a cell      WARN
  XL-CLIP         long text clipped by an occupied neighbor (wrap off)        WARN  (heuristic)

Exit: 1 if any ERROR (or any WARN with --strict). 0 otherwise. 2 on usage error.

Usage:
    py xlsx_lint.py book.xlsx [--json report.json] [--strict] [--quiet]
"""
import argparse
import json
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CFG = {
    "block_min_cells": 4,        # a "table block" must be at least this many cells
    "block_fill_ratio": 0.6,     # bbox must be at least this solid to count as a block
    "numeric_majority": 0.6,     # column is "numeric" when >= this fraction are numbers
    "clip_slack": 1.15,          # text len must exceed col width * this to call it clipped
    "scan_cap": 300,             # max rows/cols scanned per sheet (perf guard)
}

ERROR_RE = re.compile(r"^#(REF|DIV/0|VALUE|NAME\?|NAME|N/A|NULL|NUM|SPILL|CALC|GETTING_DATA)[!?]?$")
REF_IN_FORMULA = re.compile(r"#REF!")
MD_PATTERNS = [
    (re.compile(r"\*\*[^*]+\*\*"), "literal **bold** markdown"),
    (re.compile(r"__[^_]+__"), "literal __bold__ markdown"),
    (re.compile(r"^\s{0,3}#{1,6}\s"), "literal #-heading markdown"),
    (re.compile(r"^\s{0,3}[-*+]\s+\S"), "literal dash/asterisk bullet"),
    (re.compile(r"^\s*\|.+\|.+\|\s*$"), "literal pipe-table markdown"),
    (re.compile(r"\[[^\]]+\]\([^)]+\)"), "literal [link](url) markdown"),
]
PLACEHOLDER_TEXT = re.compile(
    r"click to add|lorem ipsum|your text here|insert .* here|\[?placeholder\]?|"
    r"sample text|todo:|tbd\b|xxx+|fixme",
    re.I,
)
NUMERIC_RE = re.compile(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$|^-?\d+(\.\d+)?$")


def is_number_text(v):
    """A string that clearly means a number (so storing it as text is a defect).
    Conservative: skip leading-zero codes (IDs, zips) and over-long digit strings."""
    if not isinstance(v, str):
        return False
    s = v.strip()
    if not NUMERIC_RE.match(s):
        return False
    digits = s.lstrip("-").replace(",", "").split(".")[0]
    if len(digits) > 1 and digits.startswith("0"):
        return False  # "007", "012345" -> probably an intentional code, leave it
    if len(digits) >= 15:
        return False  # long IDs/account numbers are legitimately text
    return True


def cell_text(c):
    v = c.value
    return v if isinstance(v, str) else None


# ---------- block detection (connected components of non-empty cells) ----------
def find_blocks(ws, max_r, max_c):
    nonempty = set()
    for r in range(1, max_r + 1):
        for cc in range(1, max_c + 1):
            v = ws.cell(row=r, column=cc).value
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                nonempty.add((r, cc))
    seen = set()
    blocks = []
    for cell in nonempty:
        if cell in seen:
            continue
        stack = [cell]
        comp = []
        seen.add(cell)
        while stack:
            r, cc = stack.pop()
            comp.append((r, cc))
            for dr, dc in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                nb = (r + dr, cc + dc)
                if nb in nonempty and nb not in seen:
                    seen.add(nb)
                    stack.append(nb)
        if len(comp) >= CFG["block_min_cells"]:
            rs = [p[0] for p in comp]
            csv = [p[1] for p in comp]
            bbox = (min(rs), min(csv), max(rs), max(csv))
            h = bbox[2] - bbox[0] + 1
            w = bbox[3] - bbox[1] + 1
            if (len(comp) / (h * w)) >= CFG["block_fill_ratio"] and h >= 2 and w >= 2:
                blocks.append(bbox)
    return blocks


def has_edge(cell, side):
    b = cell.border
    seg = getattr(b, side)
    return seg is not None and seg.style is not None


# ---------- linter ----------
class Linter:
    def __init__(self, cfg=CFG):
        self.cfg = cfg
        self.findings = []

    def add(self, sev, cid, sheet, ref, msg):
        self.findings.append({"id": cid, "severity": sev, "sheet": sheet,
                              "ref": ref, "message": msg})

    def run(self, path):
        wb = load_workbook(path, data_only=False)
        for ws in wb.worksheets:
            max_r = min(ws.max_row or 1, self.cfg["scan_cap"])
            max_c = min(ws.max_column or 1, self.cfg["scan_cap"])
            self.check_cells(ws, max_r, max_c)
            self.check_merges(ws)
            self.check_blocks(ws, max_r, max_c)
        return {
            "file": os.path.abspath(path),
            "summary": self.summary(),
            "findings": sorted(self.findings, key=lambda f: (f["sheet"], f["id"], f["ref"])),
        }

    def summary(self):
        s = {"ERROR": 0, "WARN": 0, "INFO": 0}
        for f in self.findings:
            s[f["severity"]] += 1
        return s

    # XL-FORMULA-ERR, XL-ARTIFACT, XL-CLIP (per-cell)
    def check_cells(self, ws, max_r, max_c):
        for r in range(1, max_r + 1):
            for cc in range(1, max_c + 1):
                c = ws.cell(row=r, column=cc)
                v = c.value
                ref = f"{get_column_letter(cc)}{r}"
                # formula / value error
                if c.data_type == "e" or (isinstance(v, str) and ERROR_RE.match(v.strip())):
                    self.add("ERROR", "XL-FORMULA-ERR", ws.title, ref,
                             f"Cell holds an error value ({str(v).strip()}).")
                elif isinstance(v, str) and v.startswith("=") and REF_IN_FORMULA.search(v):
                    self.add("ERROR", "XL-FORMULA-ERR", ws.title, ref,
                             f"Formula contains a broken #REF! reference: {v[:40]!r}.")
                # artifact text
                t = cell_text(c)
                if t:
                    for pat, why in MD_PATTERNS:
                        if pat.search(t):
                            self.add("WARN", "XL-ARTIFACT", ws.title, ref,
                                     f"{why}: {t.strip()[:40]!r}")
                            break
                    if PLACEHOLDER_TEXT.search(t):
                        self.add("WARN", "XL-ARTIFACT", ws.title, ref,
                                 f"Placeholder/boilerplate: {t.strip()[:40]!r}")
                    # clip by occupied right neighbor (wrap off)
                    self.check_clip(ws, c, r, cc, t, max_c)

    def check_clip(self, ws, c, r, cc, t, max_c):
        try:
            wrap = bool(c.alignment.wrap_text)
        except Exception:
            wrap = False
        if wrap or len(t) <= 8 or cc >= max_c:
            return
        right = ws.cell(row=r, column=cc + 1).value
        if right is None or (isinstance(right, str) and right.strip() == ""):
            return  # spills into empty cell = normal Excel, not clipped
        col = get_column_letter(cc)
        width = None
        if col in ws.column_dimensions and ws.column_dimensions[col].width:
            width = ws.column_dimensions[col].width
        if width is None:
            width = 8.43
        if len(t) > width * self.cfg["clip_slack"]:
            self.add("WARN", "XL-CLIP", ws.title, f"{col}{r}",
                     f"Text ~{len(t)} chars clipped by the occupied next cell "
                     f"(column width ~{width:.0f}; wrap off).")

    # XL-MERGE-OVERLAP
    def check_merges(self, ws):
        ranges = list(ws.merged_cells.ranges)
        for i in range(len(ranges)):
            for j in range(i + 1, len(ranges)):
                a, b = ranges[i], ranges[j]
                if (a.min_row <= b.max_row and b.min_row <= a.max_row and
                        a.min_col <= b.max_col and b.min_col <= a.max_col):
                    self.add("ERROR", "XL-MERGE-OVERLAP", ws.title,
                             f"{a.coord} & {b.coord}", "Merged ranges overlap.")

    # XL-BORDER-GAP (partial outline) + XL-NUM-AS-TEXT (per data block)
    def check_blocks(self, ws, max_r, max_c):
        for (r0, c0, r1, c1) in find_blocks(ws, max_r, max_c):
            self.check_border_gap(ws, r0, c0, r1, c1)
            self.check_num_as_text(ws, r0, c0, r1, c1)

    def check_border_gap(self, ws, r0, c0, r1, c1):
        edges = {
            "top": [ws.cell(row=r0, column=c) for c in range(c0, c1 + 1)],
            "bottom": [ws.cell(row=r1, column=c) for c in range(c0, c1 + 1)],
            "left": [ws.cell(row=r, column=c0) for r in range(r0, r1 + 1)],
            "right": [ws.cell(row=r, column=c1) for r in range(r0, r1 + 1)],
        }
        side_of = {"top": "top", "bottom": "bottom", "left": "left", "right": "right"}
        for name, cells in edges.items():
            n = len(cells)
            got = sum(1 for c in cells if has_edge(c, side_of[name]))
            if 0 < got < n:  # PARTIAL edge = the "border on some cells only" artifact
                self.add("WARN", "XL-BORDER-GAP", ws.title,
                         f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}",
                         f"{name} border present on {got}/{n} cells — incomplete outline.")

    def check_num_as_text(self, ws, r0, c0, r1, c1):
        for c in range(c0, c1 + 1):
            cells = [ws.cell(row=r, column=c) for r in range(r0, r1 + 1)]
            num = sum(1 for x in cells if isinstance(x.value, (int, float)) and not isinstance(x.value, bool))
            if num < 2 or num / len(cells) < self.cfg["numeric_majority"]:
                continue
            for x in cells:
                if is_number_text(x.value):
                    self.add("WARN", "XL-NUM-AS-TEXT", ws.title,
                             f"{get_column_letter(c)}{x.row}",
                             f"Number stored as text ({x.value!r}) in a numeric column "
                             f"— breaks math/sort.")


# ---------- reporting ----------
SEV_ORDER = {"ERROR": 0, "WARN": 1, "INFO": 2}
SEV_MARK = {"ERROR": "x", "WARN": "!", "INFO": "i"}


def print_human(report):
    s = report["summary"]
    print(f"\nXLSX LINT  {os.path.basename(report['file'])}")
    print(f"  {s['ERROR']} ERROR   {s['WARN']} WARN   {s['INFO']} INFO\n")
    if not report["findings"]:
        print("  clean -- no objective defects found.\n")
        return
    last = None
    for f in sorted(report["findings"],
                    key=lambda x: (x["sheet"], SEV_ORDER[x["severity"]], x["id"], x["ref"])):
        if f["sheet"] != last:
            print(f"-- sheet '{f['sheet']}' --")
            last = f["sheet"]
        print(f"  [{SEV_MARK[f['severity']]}] {f['id']} [{f['ref']}]: {f['message']}")
    print()


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--json", default=None)
    ap.add_argument("--strict", action="store_true")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        print("ERROR: no such file:", args.xlsx, file=sys.stderr)
        sys.exit(2)

    report = Linter().run(args.xlsx)
    if args.json:
        with open(args.json, "w", encoding="utf-8") as fh:
            json.dump(report, fh, ensure_ascii=False, indent=2)
        print("wrote", args.json)
    if not args.quiet:
        print_human(report)

    s = report["summary"]
    sys.exit(1 if s["ERROR"] > 0 or (args.strict and s["WARN"] > 0) else 0)


if __name__ == "__main__":
    main()
