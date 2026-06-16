"""
make_test_book.py — build a deliberately-flawed .xlsx to validate xlsx_lint.py.

Injects known OBJECTIVE defects (and clean controls that must NOT fire).
Usage: py make_test_book.py [out.xlsx]   (default ./_test_flawed_book.xlsx)
"""
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
THIN = Side(style="thin")
FULL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_TOP = Border(left=THIN, right=THIN, top=None, bottom=THIN)


def build(out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    # --- Block 1: CONTROL, fully bordered numeric block A1:C3 -> NO finding ---
    ctrl = [["Q", "n", "m"], [1, 10, 100], [2, 20, 200]]
    for r, row in enumerate(ctrl, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = FULL

    # --- Block 2: PARTIAL top border E1:G3 -> XL-BORDER-GAP (top 2/3) ---
    blk2 = [["Q", "n", "m"], [3, 30, 300], [4, 40, 400]]
    for r, row in enumerate(blk2, start=1):
        for c, val in enumerate(row, start=5):     # cols E,F,G
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = FULL if not (r == 1 and c == 6) else NO_TOP  # drop top on F1

    # --- Block 3: numeric column with one TEXT number A6:B9 -> XL-NUM-AS-TEXT ---
    blk3 = [["alpha", 10], ["bravo", 20], ["charlie", "30"], ["delta", 40]]  # "30" is text
    for r, row in enumerate(blk3, start=6):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # --- Scattered objective defects (caught per-cell, no block needed) ---
    ws["A12"] = "**bold left in a cell**"          # XL-ARTIFACT (markdown)
    ws["A13"] = "TODO: fill this in"               # XL-ARTIFACT (placeholder)
    err = ws["A15"]
    err.value = "#DIV/0!"
    err.data_type = "e"                            # XL-FORMULA-ERR (error value)
    ws["A16"] = "=SUM(#REF!)"                      # XL-FORMULA-ERR (broken ref)

    # --- Clip: long text with an occupied right neighbor, wrap off -> XL-CLIP ---
    a18 = ws["A18"]
    a18.value = "This label is far longer than the narrow column allows"
    a18.alignment = Alignment(wrap_text=False)
    ws["B18"] = "x"                                # neighbor occupied -> text is clipped

    # --- Overlapping merges -> XL-MERGE-OVERLAP (best effort) ---
    try:
        ws.merge_cells("D12:E13")
        ws.merge_cells("E13:F14")                  # overlaps at E13
    except Exception as e:
        print("note: could not create overlapping merges:", e)

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "_test_flawed_book.xlsx")
    print("wrote", build(out))
